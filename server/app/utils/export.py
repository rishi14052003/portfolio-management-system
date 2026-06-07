"""Utilities for exporting data"""
import csv
import io
from datetime import datetime
from typing import List, Dict
from app.models import Lead


def export_leads_to_csv(leads: List[Lead]) -> str:
    """Export leads to CSV format"""
    if not leads:
        return ""
    
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[
            'ID',
            'Full Name',
            'Phone',
            'Email',
            'Purpose',
            'Company',
            'Budget',
            'Found Via',
            'Message',
            'Submitted',
            'PDF File'
        ]
    )
    
    writer.writeheader()
    for lead in leads:
        writer.writerow({
            'ID': lead.id,
            'Full Name': lead.full_name,
            'Phone': lead.phone,
            'Email': lead.email,
            'Purpose': lead.purpose or '-',
            'Company': lead.company or '-',
            'Budget': lead.budget or '-',
            'Found Via': lead.source or '-',
            'Message': lead.introduction[:100] + '...' if len(lead.introduction) > 100 else lead.introduction,
            'Submitted': lead.created_at,
            'PDF File': lead.pdf_filename or '-'
        })
    
    return output.getvalue()


def generate_csv_filename() -> str:
    """Generate CSV filename with timestamp"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f'leads_export_{timestamp}.csv'


def export_leads_to_excel(leads: List[Lead]) -> bytes:
    """Export leads to Excel format (.xlsx)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Define headers
    headers = [
        'ID',
        'Full Name',
        'Phone',
        'Email',
        'Purpose',
        'Company',
        'Budget',
        'Found Via',
        'Message',
        'Submitted',
        'PDF File'
    ]
    
    # Add headers with styling
    header_fill = PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data rows
    for row_idx, lead in enumerate(leads, 2):
        ws.cell(row=row_idx, column=1).value = lead.id
        ws.cell(row=row_idx, column=2).value = lead.full_name
        ws.cell(row=row_idx, column=3).value = lead.phone
        ws.cell(row=row_idx, column=4).value = lead.email
        ws.cell(row=row_idx, column=5).value = lead.purpose or '-'
        ws.cell(row=row_idx, column=6).value = lead.company or '-'
        ws.cell(row=row_idx, column=7).value = lead.budget or '-'
        ws.cell(row=row_idx, column=8).value = lead.source or '-'
        ws.cell(row=row_idx, column=9).value = lead.introduction[:100] + '...' if len(lead.introduction) > 100 else lead.introduction
        ws.cell(row=row_idx, column=10).value = lead.created_at
        ws.cell(row=row_idx, column=11).value = lead.pdf_filename or '-'
    
    # Adjust column widths
    column_widths = [10, 15, 12, 20, 18, 18, 15, 15, 30, 16, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_excel_filename() -> str:
    """Generate Excel filename with timestamp"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f'leads_export_{timestamp}.xlsx'
